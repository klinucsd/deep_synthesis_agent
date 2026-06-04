"""
comprehensive_report_v2 — render the per-submission v2 PDF.

This module is the pair file for the wildfire-comprehensive-report-v2 skill.

Public entry point:
    generate_report_v2(ws_root, *, science_xlsx=None, practitioner_xlsx=None,
                       output_path=None, force=False) -> Path

The v2 PDF integrates three voices:
  1. The team's self-report PDF (referenced on the cover; not inlined).
  2. The Review Agent's rubric review (science + practitioner JSON).
  3. Human evaluator commentary (parsed from the two Preliminary-Extract xlsx
     files, grouped per rubric question).

The PDF lands at:
    <ws_root>/_comprehensive_report_v2_<team_id>_<scenario>.pdf
"""

from __future__ import annotations

import json
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Optional

from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
)

# TEAMS table lives in a sibling file for clarity.
from teams import TEAMS, team_for, parse_workspace_folder


# ---------------------------------------------------------------------------
# Rubric schema accessor — imports from the wildfire-rubric-review skill.
# Both skills are expected to be installed side-by-side under .../skills/.

_THIS_DIR = Path(__file__).resolve().parent
_RUBRIC_SKILL_DIR = _THIS_DIR.parent / "wildfire-rubric-review"
if _RUBRIC_SKILL_DIR.is_dir() and str(_RUBRIC_SKILL_DIR) not in sys.path:
    sys.path.insert(0, str(_RUBRIC_SKILL_DIR))

from wildfire_rubric import get_schema as _get_schema  # noqa: E402


def get_schema(rubric_type: str):
    return _get_schema(rubric_type)


def qtext(qspec: dict) -> str:
    return (qspec.get("text") or qspec.get("question") or qspec.get("prompt") or "").strip()


def format_answer(qspec: dict, answer: dict | None) -> tuple[str, str]:
    if answer is None:
        return ("", "")
    qtype = qspec.get("type", "single")
    if qtype == "multi" or qtype == "multi_select":
        sel = answer.get("selected") or []
        value = ", ".join(str(s) for s in sel) if sel else ""
    else:
        value = str(answer.get("value", "") or "")
    notes = str(answer.get("notes", "") or "")
    return (value, notes)


# ---------------------------------------------------------------------------
# Evaluator xlsx parsing

@dataclass
class EvaluatorResponse:
    kind: str                                    # "science" or "practitioner"
    org: str
    model: str
    evaluator_id: str
    answers: dict[int, object] = field(default_factory=dict)
    reviewed_both: bool = False
    matched_team_id: Optional[str] = None


def _norm(s: str | None) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _load_xlsx_responses(path: Path, kind: str) -> tuple[list[str], list[EvaluatorResponse]]:
    """Load one evaluator xlsx file into (headers, list[EvaluatorResponse])."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    # Per-file the sheet is named "survey_0".
    sheet_name = "survey_0" if "survey_0" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ([], [])
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if r[0] is None and (len(r) < 2 or r[1] is None) and (len(r) < 3 or r[2] is None):
            continue
        answers = {i: v for i, v in enumerate(r) if v not in (None, "")}
        org = str(r[0]).strip() if r[0] else ""
        model = str(r[1]).strip() if len(r) > 1 and r[1] else ""
        evid = str(r[2]).strip() if len(r) > 2 and r[2] else ""
        reviewed_both = (str(r[3]).strip().lower() == "yes") if len(r) > 3 and r[3] else False
        out.append(EvaluatorResponse(
            kind=kind, org=org, model=model, evaluator_id=evid,
            answers=answers, reviewed_both=reviewed_both,
        ))
    return headers, out


def _resolve_team_id(resp: EvaluatorResponse) -> Optional[str]:
    """Match an evaluator row to a TEAMS entry by org alias and (optionally) model filter."""
    org_norm = _norm(resp.org)
    model_norm = _norm(resp.model)
    candidates = []
    for t in TEAMS:
        for alias in t["evaluator_orgs"]:
            if _norm(alias) == org_norm:
                filt = t.get("evaluator_model_filter")
                if filt is None or _norm(filt) in model_norm:
                    candidates.append(t["team_id"])
                break
    return candidates[0] if candidates else None


# ---------------------------------------------------------------------------
# Column → qid mapping (fuzzy match against rubric schema text)

def _build_text_to_qid(rubric_kind: str) -> dict[str, str]:
    return {_norm(qtext(q)): q["id"] for q in get_schema(rubric_kind) if qtext(q)}


def _best_qid(target: str, text2qid: dict[str, str], min_ratio: float) -> Optional[str]:
    target_n = _norm(target)
    best_qid, best_r = None, 0.0
    for ct, qid in text2qid.items():
        if not ct:
            continue
        r = SequenceMatcher(None, target_n, ct).ratio()
        if r > best_r:
            best_qid, best_r = qid, r
    return best_qid if best_r >= min_ratio else None


def _build_column_map(headers: list[str], rubric_kind: str, min_ratio: float = 0.40) -> dict[int, tuple[str, str]]:
    """Map each xlsx column index → (qid, role). role is 'main' or 'other'."""
    text2qid = _build_text_to_qid(rubric_kind)
    out: dict[int, tuple[str, str]] = {}
    last_qid = None
    for i, hdr in enumerate(headers):
        if i < 3 or not hdr:
            continue
        h = _norm(hdr)
        if h.startswith("other - "):
            sub = h[8:]
            qid = _best_qid(sub, text2qid, min_ratio) or last_qid
            if qid:
                out[i] = (qid, "other")
            continue
        qid = _best_qid(hdr, text2qid, min_ratio)
        if qid:
            out[i] = (qid, "main")
            last_qid = qid
    return out


def _cells_for_qid(answers: dict[int, object], col_map: dict[int, tuple[str, str]], qid: str) -> tuple[Optional[str], Optional[str]]:
    main_value = None
    other_value = None
    for col_idx, (cqid, role) in col_map.items():
        if cqid != qid:
            continue
        v = answers.get(col_idx)
        if v in (None, ""):
            continue
        if role == "main":
            main_value = str(v).strip()
        elif role == "other":
            other_value = str(v).strip()
    return (main_value, other_value)


# ---------------------------------------------------------------------------
# Styles

_SLATE = colors.HexColor("#1f2937")
_SUBTLE = colors.HexColor("#4b5563")
_RULE = colors.HexColor("#d1d5db")
_BG_LIGHT = colors.HexColor("#f9fafb")
_ACCENT_EVAL = colors.HexColor("#1e3a8a")
_ACCENT_AGENT = colors.HexColor("#0f766e")
_BANNER_TEXT = colors.HexColor("#92400e")

_BASE = getSampleStyleSheet()

_TITLE = ParagraphStyle("Title", parent=_BASE["Title"],
    fontName="Helvetica-Bold", fontSize=20, leading=24,
    textColor=_SLATE, alignment=TA_LEFT, spaceBefore=0, spaceAfter=6)
_SUBTITLE = ParagraphStyle("Subtitle", parent=_BASE["BodyText"],
    fontName="Helvetica", fontSize=13, leading=17,
    textColor=_SUBTLE, spaceBefore=0, spaceAfter=2)
_SECTION_EVAL = ParagraphStyle("SectionEval", parent=_BASE["Heading1"],
    fontName="Helvetica-Bold", fontSize=15, leading=19,
    textColor=_ACCENT_EVAL, spaceBefore=16, spaceAfter=8)
_SECTION_AGENT = ParagraphStyle("SectionAgent", parent=_BASE["Heading1"],
    fontName="Helvetica-Bold", fontSize=15, leading=19,
    textColor=_ACCENT_AGENT, spaceBefore=16, spaceAfter=8)
_H2 = ParagraphStyle("H2", parent=_BASE["Heading2"],
    fontName="Helvetica-Bold", fontSize=11.5, leading=15,
    textColor=_SLATE, spaceBefore=10, spaceAfter=4)
_QUESTION = ParagraphStyle("Question", parent=_BASE["BodyText"],
    fontName="Helvetica-Bold", fontSize=10.5, leading=14,
    textColor=_SLATE, spaceBefore=10, spaceAfter=3)
_LABEL = ParagraphStyle("Label", parent=_BASE["BodyText"],
    fontName="Helvetica-Bold", fontSize=9.5, leading=12.5,
    textColor=_SLATE, spaceBefore=0, spaceAfter=0)
_BODY = ParagraphStyle("Body", parent=_BASE["BodyText"],
    fontName="Helvetica", fontSize=9.5, leading=12.5,
    textColor=_SLATE, alignment=TA_LEFT, spaceBefore=0, spaceAfter=2)
_NOTES = ParagraphStyle("Notes", parent=_BASE["BodyText"],
    fontName="Helvetica-Oblique", fontSize=9, leading=12,
    textColor=_SUBTLE, spaceBefore=0, spaceAfter=2)
_EVAL_LINE = ParagraphStyle("EvalLine", parent=_BASE["BodyText"],
    fontName="Helvetica", fontSize=9.5, leading=13,
    textColor=_SLATE, leftIndent=14, spaceBefore=1, spaceAfter=1)
_BANNER = ParagraphStyle("Banner", parent=_BASE["BodyText"],
    fontName="Helvetica-Bold", fontSize=10, leading=13,
    textColor=_BANNER_TEXT, alignment=TA_LEFT, spaceBefore=4, spaceAfter=4)
_NO_ENTRY = ParagraphStyle("NoEntry", parent=_BASE["BodyText"],
    fontName="Helvetica-Oblique", fontSize=9, leading=12,
    textColor=colors.HexColor("#9ca3af"), spaceBefore=0, spaceAfter=2)


_COMMA_NO_SPACE = re.compile(r",(?!\s)")


def _esc(text: object) -> str:
    if text is None:
        return ""
    s = str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    s = s.replace("\n", "<br/>")
    s = _COMMA_NO_SPACE.sub(", ", s)
    return s


# ---------------------------------------------------------------------------
# Section builders

def _find_team_self_report(ws_root: Path) -> Optional[str]:
    """Return the filename of any PDF in the workspace that looks like a team self-report."""
    for p in ws_root.iterdir():
        if p.is_file() and p.suffix.lower() == ".pdf" and not p.name.startswith("_comprehensive_report"):
            return p.name
    return None


def _cover(team: dict, scenario: str, ws_root: Path, evaluators: list[EvaluatorResponse]) -> list:
    flow = []
    flow.append(Paragraph("Comprehensive Submission Review", _TITLE))
    flow.append(Paragraph(
        f"<b>{_esc(team['tracking_org'])}</b> &mdash; {_esc(team['tracking_model'])}",
        _SUBTITLE,
    ))
    flow.append(Paragraph(
        f"Scenario: <b>{_esc(scenario)}</b>"
        f"{' &nbsp;·&nbsp; (combined-report team)' if team.get('combined_report') else ''}",
        _SUBTITLE,
    ))
    flow.append(Paragraph(
        f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        _SUBTITLE,
    ))
    flow.append(Spacer(1, 0.18 * inch))

    n_sci = sum(1 for r in evaluators if r.kind == "science")
    n_prac = sum(1 for r in evaluators if r.kind == "practitioner")
    eval_summary = f"{n_sci} Science evaluator(s); {n_prac} Practitioner evaluator(s)"

    pdf_filename = _find_team_self_report(ws_root)
    src_rows = [
        ["Team self-report:", pdf_filename or "(none found in workspace)"],
        ["Evaluator responses:", eval_summary],
    ]
    src_table = Table(
        [[Paragraph(label, _LABEL), Paragraph(_esc(val), _BODY)] for label, val in src_rows],
        colWidths=[1.9 * inch, 4.7 * inch],
    )
    src_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BACKGROUND", (0, 0), (-1, -1), _BG_LIGHT),
        ("BOX", (0, 0), (-1, -1), 0.4, _RULE),
        ("INNERGRID", (0, 0), (-1, -1), 0.2, _RULE),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    flow.append(src_table)
    return flow


_AT_A_GLANCE_SCI = [
    ("q1", "Model name"),
    ("q4", "Static vs dynamic"),
    ("q4a", "Synoptic weather applied"),
    ("q4b", "Weather variables"),
    ("q9", "Ember dispersion"),
    ("q11", "Historical occurrences"),
    ("q13", "Population demographics"),
    ("q15", "Topography accounted"),
    ("q17", "Built environment"),
    ("q24", "Vegetation accounted"),
]
_AT_A_GLANCE_PRAC = [
    ("q4", "Primary appropriate use"),
    ("q6", "Secondary appropriate use"),
    ("qf1", "Confidence in usability"),
]


def _at_a_glance(rubric_sci: dict, rubric_prac: dict) -> list:
    flow = [Paragraph("Model at a glance", _SECTION_AGENT)]
    flow.append(Paragraph(
        "<i>Compact summary pulled from the Review Agent's rubric review. "
        "Full per-question detail appears in the rubric sections below.</i>",
        _NOTES,
    ))
    rows = [["Attribute", "Value"]]

    def add(rubric_json, schema_kind, qid, label):
        if not rubric_json:
            return
        ans = (rubric_json.get("answers") or {}).get(qid)
        if not ans:
            return
        qspec = next((q for q in get_schema(schema_kind) if q["id"] == qid), {"type": "single"})
        value, _notes = format_answer(qspec, ans)
        if value:
            rows.append([label, value[:200]])

    for qid, label in _AT_A_GLANCE_SCI:
        add(rubric_sci, "science", qid, label)
    for qid, label in _AT_A_GLANCE_PRAC:
        add(rubric_prac, "practitioner", qid, label)

    if len(rows) <= 1:
        return []

    paragraphed = [[Paragraph(_esc(c), _BODY) for c in r] for r in rows]
    t = Table(paragraphed, colWidths=[2.0 * inch, 4.6 * inch])
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BACKGROUND", (0, 0), (-1, 0), _BG_LIGHT),
        ("TEXTCOLOR", (0, 0), (-1, 0), _SLATE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("LINEBELOW", (0, 0), (-1, 0), 0.6, _RULE),
        ("BOX", (0, 0), (-1, -1), 0.4, _RULE),
        ("INNERGRID", (0, 0), (-1, -1), 0.2, _RULE),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    flow.append(t)
    return flow


def _evaluator_section_per_q(evaluators: list[EvaluatorResponse], headers: list[str],
                              rubric_kind: str, section_title: str) -> list:
    flow = [Paragraph(section_title, _SECTION_EVAL)]
    rs = [r for r in evaluators if r.kind == rubric_kind]
    if not rs:
        flow.append(Paragraph(
            f"No {rubric_kind} evaluator responses recorded for this submission.", _BODY,
        ))
        return flow
    flow.append(Paragraph(
        f"<i>{len(rs)} {rubric_kind} evaluator(s) contributed responses. "
        "For each rubric question below, every evaluator's answer is listed.</i>",
        _NOTES,
    ))
    col_map = _build_column_map(headers, rubric_kind)
    schema = get_schema(rubric_kind)
    for qspec in schema:
        qid = qspec["id"]
        lines = []
        for r in sorted(rs, key=lambda x: x.evaluator_id):
            main, other = _cells_for_qid(r.answers, col_map, qid)
            if not main and not other:
                continue
            parts = []
            if main:
                parts.append(_esc(main))
            if other:
                parts.append(f"<font color='#6b7280'><i>Other: {_esc(other)}</i></font>")
            lines.append(
                f"<b>{_esc(r.evaluator_id)}:</b> " + " &nbsp;·&nbsp; ".join(parts)
            )
        if not lines:
            continue
        flow.append(Paragraph(
            f"<b>{_esc(qid)}.</b> {_esc(qtext(qspec))}", _QUESTION,
        ))
        for line in lines:
            flow.append(Paragraph(line, _EVAL_LINE))
    return flow


def _agent_rubric_section(rubric_json: dict, rubric_kind: str, title: str) -> list:
    flow = [Paragraph(title, _SECTION_AGENT)]
    answers = rubric_json.get("answers") or {}
    branching = rubric_json.get("branching_notes") or {}
    gaps = rubric_json.get("evidence_gaps") or {}
    team_only = rubric_json.get("team_only") or {}

    for qspec in get_schema(rubric_kind):
        qid = qspec["id"]
        ans = answers.get(qid)
        flow.append(Paragraph(
            f"<b>{_esc(qid)}.</b> {_esc(qtext(qspec))}", _QUESTION,
        ))
        if ans is not None:
            value, notes = format_answer(qspec, ans)
            if value:
                flow.append(Paragraph(f"<b>Answer:</b> {_esc(value)}", _BODY))
            if notes:
                flow.append(Paragraph(f"<b>Details:</b> {_esc(notes)}", _NOTES))
        elif qid in branching:
            flow.append(Paragraph(
                f"<b>Skipped (branching):</b> {_esc(branching[qid])}", _NOTES,
            ))
        elif qid in gaps:
            flow.append(Paragraph(
                f"<b>Evidence gap:</b> {_esc(gaps[qid])}", _NOTES,
            ))
        elif qid in team_only:
            flow.append(Paragraph(
                f"<b>Team-only:</b> {_esc(team_only[qid])} <i>(subjective rating; not extracted)</i>",
                _NOTES,
            ))
        else:
            flow.append(Paragraph("<i>(no entry)</i>", _NO_ENTRY))
    return flow


def _coverage_section(rubric_sci: dict, rubric_prac: dict) -> list:
    flow = [Paragraph("Coverage notes from the Review Agent", _SECTION_AGENT)]
    for label, rj, kind in (
        ("Science rubric", rubric_sci, "science"),
        ("Practitioner rubric", rubric_prac, "practitioner"),
    ):
        if not rj:
            continue
        flow.append(Paragraph(label, _H2))
        any_entry = False
        for bucket_name, bucket_label in (
            ("evidence_gaps", "Evidence gaps"),
            ("branching_notes", "Branching skips"),
        ):
            bucket = rj.get(bucket_name) or {}
            if not bucket:
                continue
            any_entry = True
            flow.append(Paragraph(f"<b>{bucket_label}:</b>", _BODY))
            for qid, reason in bucket.items():
                qspec = next((q for q in get_schema(kind) if q["id"] == qid), None)
                qtxt = qtext(qspec) if qspec else qid
                flow.append(Paragraph(
                    f"&bull; <b>{_esc(qid)}</b> &nbsp; {_esc(qtxt[:130])} &mdash; <i>{_esc(reason)}</i>",
                    _BODY,
                ))
        if not any_entry:
            flow.append(Paragraph("<i>(no gaps or skips)</i>", _NOTES))
    return flow


# ---------------------------------------------------------------------------
# xlsx discovery

_XLSX_SEARCH_LOCATIONS = [
    ".",
    "..",
    "_deep_synthesis_sage_",
]
_SCIENCE_XLSX_PATTERNS = [
    "Preliminary-Extract-Science-Technical-Evaluator-Feedback*.xlsx",
    "*Science*Technical*Evaluator*.xlsx",
    "*science*evaluator*.xlsx",
]
_PRACTITIONER_XLSX_PATTERNS = [
    "Preliminary-Extract-User-Practitioner-Evaluation-Feedback*.xlsx",
    "*Practitioner*Evaluation*.xlsx",
    "*practitioner*.xlsx",
]


def _find_xlsx(start: Path, patterns: list[str]) -> Optional[Path]:
    """Walk a few likely locations for one of the patterns; return first match."""
    seen = set()
    for loc in _XLSX_SEARCH_LOCATIONS:
        d = (start / loc).resolve()
        if d in seen or not d.is_dir():
            continue
        seen.add(d)
        for pat in patterns:
            matches = sorted(d.glob(pat))
            if matches:
                return matches[0]
    return None


# ---------------------------------------------------------------------------
# Top-level entry point

def generate_report_v2(
    ws_root: str | Path,
    *,
    science_xlsx: Optional[Path] = None,
    practitioner_xlsx: Optional[Path] = None,
    output_path: Optional[Path] = None,
    force: bool = False,
) -> Path:
    """Render the v2 comprehensive submission-review PDF for one workspace."""
    ws_root = Path(ws_root).resolve()
    if not ws_root.is_dir():
        raise FileNotFoundError(f"Workspace not found: {ws_root}")

    team_id, scenario = parse_workspace_folder(ws_root.name)
    if team_id is None or scenario is None:
        raise ValueError(
            f"Workspace folder name {ws_root.name!r} does not match the expected "
            f"Deep_Synthesis__<team_id>_(Forest|Prairie) pattern, or the team_id "
            f"is not in the TEAMS table inside this skill."
        )
    team = team_for(team_id)
    assert team is not None  # team_for is consistent with parse_workspace_folder

    # Inputs: rubric JSONs (required).
    sci_path = ws_root / "_rubric_science.json"
    prac_path = ws_root / "_rubric_practitioner.json"
    missing = [p.name for p in (sci_path, prac_path) if not p.exists()]
    if missing:
        raise FileNotFoundError(
            f"Workspace {ws_root.name} is missing required rubric file(s): "
            f"{', '.join(missing)}. Run the wildfire-rubric-review skill on "
            f"this workspace first."
        )
    rubric_sci = json.loads(sci_path.read_text())
    rubric_prac = json.loads(prac_path.read_text())

    # Inputs: evaluator xlsx (auto-discover if not provided).
    if science_xlsx is None:
        science_xlsx = _find_xlsx(ws_root.parent, _SCIENCE_XLSX_PATTERNS) \
                       or _find_xlsx(Path.cwd(), _SCIENCE_XLSX_PATTERNS)
    if practitioner_xlsx is None:
        practitioner_xlsx = _find_xlsx(ws_root.parent, _PRACTITIONER_XLSX_PATTERNS) \
                            or _find_xlsx(Path.cwd(), _PRACTITIONER_XLSX_PATTERNS)
    if science_xlsx is None or practitioner_xlsx is None:
        missing_xlsx = []
        if science_xlsx is None: missing_xlsx.append("Science evaluator xlsx")
        if practitioner_xlsx is None: missing_xlsx.append("Practitioner evaluator xlsx")
        raise FileNotFoundError(
            "Could not auto-discover evaluator xlsx file(s): " + ", ".join(missing_xlsx) +
            ". Searched the workspace's parent directory and the current working directory. "
            "Pass explicit paths via science_xlsx= and practitioner_xlsx= if they live elsewhere."
        )

    sci_headers, sci_rows = _load_xlsx_responses(Path(science_xlsx), "science")
    prac_headers, prac_rows = _load_xlsx_responses(Path(practitioner_xlsx), "practitioner")
    all_responses = sci_rows + prac_rows
    for r in all_responses:
        r.matched_team_id = _resolve_team_id(r)
    evaluators = [r for r in all_responses if r.matched_team_id == team_id]

    # Output path.
    if output_path is None:
        output_path = ws_root / f"_comprehensive_report_v2_{team_id}_{scenario}.pdf"
    output_path = Path(output_path)
    if output_path.exists() and not force:
        # Caller (the SKILL.md flow) is supposed to do the skip-check, but be safe.
        # Still re-render — the caller can pass force=False from skill but we don't
        # want to silently skip if they called us directly.
        pass

    # Compose the flow.
    flow: list = []
    flow.extend(_cover(team, scenario, ws_root, evaluators))

    glance = _at_a_glance(rubric_sci, rubric_prac)
    if glance:
        flow.append(Spacer(1, 0.18 * inch))
        flow.extend(glance)

    flow.append(PageBreak())
    flow.extend(_evaluator_section_per_q(
        evaluators, sci_headers, "science", "Evaluator Commentary — Science Rubric",
    ))
    flow.append(PageBreak())
    flow.extend(_evaluator_section_per_q(
        evaluators, prac_headers, "practitioner", "Evaluator Commentary — Practitioner Rubric",
    ))

    flow.append(PageBreak())
    flow.extend(_agent_rubric_section(
        rubric_sci, "science", "Review Agent — Science Rubric Based Evaluation",
    ))
    flow.append(PageBreak())
    flow.extend(_agent_rubric_section(
        rubric_prac, "practitioner", "Review Agent — Practitioner Rubric Based Evaluation",
    ))

    cov = _coverage_section(rubric_sci, rubric_prac)
    if cov:
        flow.append(PageBreak())
        flow.extend(cov)

    doc = SimpleDocTemplate(
        str(output_path), pagesize=letter,
        leftMargin=0.85 * inch, rightMargin=0.85 * inch,
        topMargin=0.7 * inch, bottomMargin=0.7 * inch,
        title=f"Comprehensive Review — {team['tracking_org']} ({scenario})",
        author="Review Agent",
    )

    org_label = team["tracking_org"]
    def _draw_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(_SUBTLE)
        canvas.drawString(0.85 * inch, 0.4 * inch, f"{org_label} — {scenario}")
        canvas.drawRightString(letter[0] - 0.85 * inch, 0.4 * inch, f"Page {doc.page}")
        canvas.restoreState()

    doc.build(flow, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
    return output_path


# ---------------------------------------------------------------------------
# CLI

if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("ws_root", help="Path to one workspace folder")
    ap.add_argument("--science-xlsx", default=None)
    ap.add_argument("--practitioner-xlsx", default=None)
    ap.add_argument("--output-path", default=None)
    ap.add_argument("--force", action="store_true")
    args = ap.parse_args()
    path = generate_report_v2(
        args.ws_root,
        science_xlsx=Path(args.science_xlsx) if args.science_xlsx else None,
        practitioner_xlsx=Path(args.practitioner_xlsx) if args.practitioner_xlsx else None,
        output_path=Path(args.output_path) if args.output_path else None,
        force=args.force,
    )
    print(f"Wrote: {path}")
