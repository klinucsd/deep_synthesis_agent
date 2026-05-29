"""
wildfire_rubric — schemas + writer + xlsx renderer for the Wildfire Risk
Modeling Exercise's two evaluation rubrics (Science Evaluator and
Practitioner Evaluator), version 2026-05-18.

Source of truth for the schemas is the same xlsx Sage uses to render the
filled review:
    skills/wildfire-rubric-science/Exercise-Evaluation-Rubric-20260518.xlsx
(The companion `wildfire-rubric-practitioner` skill imports from this
same directory, so the module + template live together in one place.)

There is no separate "rubric per location". The May 18 rubric is filled
once per (evaluator, submission); the per-location coverage is captured
inside the rubric itself (Q3 for Science, the second Q2 for Practitioner).

This module has no LLM calls. The agent passes an `answers` dict, this
module validates the answers, writes `_review_<type>_<workspace>.json`,
and renders `_review_<type>_<workspace>_filled.xlsx` automatically from
that JSON by overlaying the blank rubric template.
"""

import json
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

# This .py file ships in skills/wildfire-rubric-science/, next to the xlsx
# template. Paths are anchored on this module's __file__ so the skill works
# regardless of where the agent sys.path.insert()s it from.
_LIB_DIR = Path(__file__).resolve().parent
RUBRIC_TEMPLATE_XLSX = _LIB_DIR / "Exercise-Evaluation-Rubric-20260518.xlsx"

SCIENCE_SHEET      = "Science Eval Rubric"
PRACTITIONER_SHEET = "Practitioner Eval Rubric"


# ---------------------------------------------------------------------------
# Rubric schemas — verbatim from the 2026-05-18 xlsx
#
# Each entry:
#   id           : question id from the rubric's "No" column. Where the
#                  rubric has periods (e.g., "7a."), the period is dropped.
#                  Web-form-only questions use ids like "qf1", "qf2"... to
#                  distinguish them from Excel-rubric ids.
#   text         : exact question wording
#   type         : "single_select" | "multi_select" | "narrative" | "text"
#   options      : list of valid option labels for single/multi (None
#                  otherwise)
#   narrative    : True when the rubric expects a free-text "notes" string
#                  (in addition to or instead of an option choice)
#   branch_when  : optional condition (dict {parent_id: value_or_predicate})
#                  for when this question applies. Sub-questions outside
#                  an applicable branch are skipped and recorded in
#                  branching_notes.
#   xlsx_row     : the 1-based row index in the source sheet (header is
#                  row 1; q1 is row 2; ...). Used by render_xlsx() to fill
#                  the right cell back into the template. None for
#                  web-form-only questions that have no row in the Excel.
#
# Optional fields (web-form-only questions):
#   source             : "science_form" or "practitioner_form" — marks the
#                        question as originating in the Google web form
#                        (not in the Excel rubric).
#   form_question_name : the canonical question name from the Survey123
#                        form (e.g., "indicate_the_model_organization") —
#                        used to join Sage's extraction with the team's
#                        web-form responses when those land.
#   team_only          : True for questions that ask the reviewer's
#                        subjective experience or a 1-5 rating. Sage does
#                        NOT attempt to extract these from submission
#                        materials; the team's web-form answer is the only
#                        source.
# ---------------------------------------------------------------------------

SCIENCE_RUBRIC = [
    {"id": "q1", "text": "Model Name",
     "type": "text", "options": None, "narrative": True, "xlsx_row": 2},

    {"id": "q2", "text": "Indicate your Evaluator ID",
     "type": "text", "options": None, "narrative": False, "xlsx_row": 3},

    {"id": "q3", "text":
        "Did you review the model outputs for both \"Town of Forests\" and "
        "\"Town of Prairies\" before completing this evaluation?",
     "type": "single_select", "options": ["Yes", "No"],
     "narrative": False, "xlsx_row": 4},

    {"id": "q4", "text":
        "Is the model output classified as \"static\" or \"dynamic\"? "
        "Select one and provide notes in the narrative.",
     "type": "single_select", "options": ["Static", "Dynamic"],
     "narrative": True, "xlsx_row": 5},

    {"id": "q4a", "text":
        "If DYNAMIC - Did the model apply the synoptic weather data provided?",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "branch_when": {"q4": "Dynamic"}, "xlsx_row": 6},

    {"id": "q4b", "text":
        "If YES to 4a - Which of the weather variables did the model "
        "incorporate? Select all that apply and provide notes in the narrative.",
     "type": "multi_select",
     "options": ["Wind Speed", "Wind Direction",
                 "Wind Gusts at Speed & Frequency", "Relative Humidity",
                 "Other"],
     "narrative": True, "branch_when": {"q4a": "Yes"}, "xlsx_row": 7},

    {"id": "q4c", "text":
        "How did the model incorporate weather variables into the model?",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q4": "Dynamic"}, "xlsx_row": 8},

    {"id": "q4d", "text":
        "Did the model incorporate additional, supplemental weather "
        "information? Select all that apply and provide notes in the narrative.",
     "type": "multi_select",
     "options": ["Full Atmospheric Profiles", "Vertical Profiles",
                 "Vertical Transects", "Height by Time Plots",
                 "Fosberg-based Daily Fire Weather Index",
                 "Canadian-based Daily Fire Weather Index", "Other"],
     "narrative": True, "branch_when": {"q4": "Dynamic"}, "xlsx_row": 9},

    {"id": "q4e", "text":
        "If STATIC - Did the model apply weather conditions? Select one "
        "option and provide notes on which weather variables and the extent "
        "by which weather was included in the narrative.",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "branch_when": {"q4": "Static"}, "xlsx_row": 10},

    {"id": "q4f", "text":
        "Does the model account for ember dispersion and/or transport? Is "
        "the ember component tied to the weather data?",
     "type": "single_select", "options": ["Yes", "No"],
     "narrative": False, "xlsx_row": 11},

    {"id": "q4g", "text":
        "If YES - Describe in the narrative the ember component of the model.",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q4f": "Yes"}, "xlsx_row": 12},

    {"id": "q5", "text": "Does the model use historical occurrence data?",
     "type": "single_select", "options": ["Yes", "No"],
     "narrative": False, "xlsx_row": 13},

    {"id": "q5a", "text":
        "If YES - What is the source of historical occurrence data? "
        "And how is it used?",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q5": "Yes"}, "xlsx_row": 14},

    {"id": "q6", "text":
        "Does the model include any social, demographic data?",
     "type": "single_select", "options": ["Yes", "No"],
     "narrative": False, "xlsx_row": 15},

    {"id": "q6a", "text":
        "If YES, What is the source of the socio-economic and/or demographic "
        "data? And how is it used?",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q6": "Yes"}, "xlsx_row": 16},

    {"id": "q7", "text":
        "Does the model include any fire protection capacity of a local fire "
        "department or other fire service within the area?",
     "type": "single_select", "options": ["Yes", "No"],
     "narrative": False, "xlsx_row": 17},

    {"id": "q7a", "text":
        "If YES, what fire protection capacity variables are included and "
        "how are they factored into the model?",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q7": "Yes"}, "xlsx_row": 18},

    {"id": "q8", "text":
        "At what scale is the model output provided at? Select only one option.",
     "type": "single_select",
     "options": ["Less than 1 Meter / Sub-Meter", "10s of Meters",
                 "100s of Meters", "1 Kilometer", "Several Kilometers",
                 "10s of Kilometers"],
     "narrative": False, "xlsx_row": 19},

    {"id": "q9", "text":
        "Which of the provided exercise data did the model output use? "
        "Select all that apply.",
     "type": "multi_select",
     "options": ["Surface Layer", "Treelist", "Voxelized Tree Data",
                 "Building Layer", "Parcel Layer", "Road Network Layer",
                 "Ignition Point", "Synoptic Weather Information"],
     "narrative": False, "xlsx_row": 20},

    {"id": "q9a", "text":
        "Did the model account for the detail on trees provided in the "
        "Treelist? And to what extent? Select one and provide notes in the "
        "narrative.",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "xlsx_row": 21},

    {"id": "q9ai", "text":
        "If YES to 9a - List which of the specific vegetation details "
        "the model incorporated.",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q9a": "Yes"}, "xlsx_row": 22},

    {"id": "q9b", "text":
        "Did the model account for topographic features in the Surface "
        "data? And to what extent? Select one and provide notes in the "
        "narrative.",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "xlsx_row": 23},

    {"id": "q9bi", "text":
        "If YES to 9b - List which of the specific surface data features "
        "the model incorporated.",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q9b": "Yes"}, "xlsx_row": 24},

    {"id": "q9bii", "text":
        "How did the inclusion of these specific surface data features "
        "factor into the models' risk score or ranking output?",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q9b": "Yes"}, "xlsx_row": 25},

    {"id": "q9biii", "text":
        "Does it include a wildfire propagation/spread model as a component "
        "part? Select one and describe in the notes.",
     "type": "single_select", "options": ["Yes", "No"],
     "narrative": True, "xlsx_row": 26},

    {"id": "q9c", "text":
        "Did the model account for building characteristics in the Building "
        "data? And to what extent? Select one and provide notes in the "
        "narrative.",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "xlsx_row": 27},

    {"id": "q9ci", "text":
        "Which of the following building variables does the model account "
        "for? Select all that apply.",
     "type": "multi_select",
     "options": ["Structural Separation Distance", "Year Built",
                 "Proximity to Fire Prone Area", "Other"],
     "narrative": True, "branch_when": {"q9c": "Yes"}, "xlsx_row": 28},

    {"id": "q9cii", "text":
        "If YES to 9c - List which of the specific building characteristics "
        "the model incorporated.",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q9c": "Yes"}, "xlsx_row": 29},

    {"id": "q9ciii", "text":
        "How did the inclusion of these building characteristics factor "
        "into the model's risk score or ranking output?",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q9c": "Yes"}, "xlsx_row": 30},

    {"id": "q9civ", "text":
        "Does the model account for probability of flame spread and/or "
        "radiative heat transfer between structures? If yes, indicate which "
        "and describe in the narrative.",
     "type": "single_select",
     "options": ["Yes - Flame Spread", "Yes - Radiative Heat Transfer",
                 "Yes - BOTH", "None"],
     "narrative": True, "xlsx_row": 31},

    {"id": "q9d", "text":
        "Indicate which ladder fuels in the built environment?",
     "type": "multi_select",
     "options": ["None", "Fences (combustible)", "Fences (non-combustible)",
                 "Detached Garage or Shed (hardened)",
                 "Detached Garage or Shed (non-hardened)", "Vehicles",
                 "Deck (at-grade, above grade)",
                 "Other Combustibles on the Parcel"],
     "narrative": False, "xlsx_row": 32},

    {"id": "q10", "text":
        "Did the model account for mitigation actions taken in the landscape "
        "in how risk is quantified?",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "xlsx_row": 33},

    {"id": "q10a", "text":
        "If YES - Indicate which of the following Landscape mitigations the "
        "model accounts for. Select all that apply and provide notes in the "
        "narrative regarding the type and scope of treatment (i.e. prescribed "
        "burn vs. mastication).",
     "type": "multi_select",
     "options": ["Fuels Reduction in Wildland", "Fuels Reduction in WUI",
                 "Parcel-level Defensible Space Treatments at 10 ft",
                 "Parcel-level Defensible Space treatments at 30ft",
                 "Parcel-level Defensible Space treatments at 100ft",
                 "Other (if other, list)"],
     "narrative": True, "branch_when": {"q10": "Yes"}, "xlsx_row": 34},

    {"id": "q10b", "text":
        "If YES - Indicate which of the following Structural features the "
        "model accounts for. Select all that apply.",
     "type": "multi_select",
     "options": ["Roof Materials", "Exterior Siding Material",
                 "Building Square Footage", "Year Built", "Eave Type",
                 "Vent Type", "Windows by type", "Other (if other, list)"],
     "narrative": True, "branch_when": {"q10": "Yes"}, "xlsx_row": 35},

    {"id": "q10c", "text":
        "If YES - Describe how the model measured variances in the structural "
        "features included within the model.",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q10": "Yes"}, "xlsx_row": 36},

    {"id": "q11", "text": "Did the model apply other, supplemental data?",
     "type": "single_select", "options": ["Yes", "No"],
     "narrative": False, "xlsx_row": 37},

    {"id": "q11a", "text": "If YES - List the other, supplemental data used.",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q11": "Yes"}, "xlsx_row": 38},

    {"id": "q11b", "text":
        "If YES - For the other, supplemental data use, describe how it "
        "was applied and whether it impacted model output values.",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q11": "Yes"}, "xlsx_row": 39},

    {"id": "q12", "text":
        "Does the model output include some kind of quantitative spatial "
        "distribution of risk (or it's equivalent)? Select all that applies "
        "and provide notes in the narrative.",
     "type": "multi_select",
     "options": ["Areas Impacted by Spread",
                 "Number of Structures Impacted", "Types of Structures",
                 "Specific Areas within a Subdivision",
                 "Specific Areas in the Wildland or WUI",
                 "Other (if other, list)"],
     "narrative": True, "xlsx_row": 40},

    {"id": "q13", "text":
        "Does the model output provide a rank, score, or value for \"Risk\"?",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": False, "xlsx_row": 41},

    {"id": "q13a", "text":
        "If YES - Indicate the type of risk ranking or score provided.",
     "type": "single_select",
     "options": ["Binary", "Numerical", "Categorical", "Other"],
     "narrative": True, "branch_when": {"q13": "Yes"}, "xlsx_row": 42,
     "_option_descriptions": {
         "Binary":      "Classify areas as either high risk or not. The risk score indicates whether an area exceeds a threshold.",
         "Numerical":   "Continuous risk scores that can be visualized in quantized buckets, even buckets, custom buckets.",
         "Categorical": "Assign discrete risk categories to areas. The risk score is a numeric identifier that maps to a category.",
     }},

    {"id": "q13b", "text":
        "If YES - Does the risk ranking or score include economic variables "
        "such as expected financial loss? If Yes, provide notes in the narrative.",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "branch_when": {"q13": "Yes"}, "xlsx_row": 43},

    {"id": "q13c", "text":
        "Describe the risk ranking or score provided by the model based on "
        "your assessment.",
     "type": "narrative", "options": None, "narrative": True, "xlsx_row": 44},

    {"id": "q14", "text":
        "Were there any risk scores, rankings, or values that did not appear "
        "to align with the characteristics of the area? Select one option "
        "and provide notes in the narrative.",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "xlsx_row": 45},

    {"id": "q15", "text":
        "Were there any gaps in the risk score, rankings, or values for "
        "certain areas? Select one option and provide notes in the narrative.",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "xlsx_row": 46},

    {"id": "q16", "text":
        "Does the model differentiate between environment types/classifications "
        "for wildland, wildland urban interface, wildland urban intermix, "
        "urban/suburban area?",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": False, "xlsx_row": 47},

    {"id": "q16a", "text":
        "If YES - Describe how the model accounts for the distinction in "
        "these environment types and how they are defined and measured in "
        "the area.",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q16": "Yes"}, "xlsx_row": 48},

    {"id": "q16b", "text":
        "If NO - Describe what environment type is applied to the area "
        "exclusively.",
     "type": "single_select",
     "options": ["Wildland", "Wildland Urban Interface",
                 "Wildland Urban Intermix", "Urban/Suburban", "Other"],
     "narrative": True, "branch_when": {"q16": "No"}, "xlsx_row": 49},

    {"id": "q17", "text":
        "Is the model output a full measure or quantification of \"Risk\"?",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": False, "xlsx_row": 50},

    {"id": "q17a", "text":
        "If YES - Explain why and describe the severity or consequence variable.",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q17": "Yes"}, "xlsx_row": 51},

    {"id": "q17b", "text":
        "If NO - Indicate which of the following the model measures. "
        "Select all that apply.",
     "type": "multi_select",
     "options": ["Hazard Exposure", "Vulnerability", "Likelihood",
                 "Burn Probability", "Other"],
     "narrative": True, "branch_when": {"q17": "No"}, "xlsx_row": 52},

    {"id": "q18", "text":
        "Which of the following approaches does the model apply? Select "
        "all that apply.",
     "type": "multi_select",
     "options": ["Empirical/Statistical", "Semi-Physics", "Reduced Physics",
                 "ML / LLM", "Retrieval Augmented (RAG) Model", "AI Agent",
                 "Agentic AI", "Other"],
     "narrative": True, "xlsx_row": 53},

    {"id": "q18a", "text": "Provide any relevant notes on the approach.",
     "type": "narrative", "options": None, "narrative": True, "xlsx_row": 54},

    {"id": "q18b", "text":
        "For models that include AI methods (ML/LLM, RAG, AI Agent, Agentic), "
        "describe how they are incorporated into the model.",
     "type": "narrative", "options": None, "narrative": True, "xlsx_row": 55},

    {"id": "q19", "text": "Describe the model's top 3 strengths.",
     "type": "narrative", "options": None, "narrative": True, "xlsx_row": 56},

    {"id": "q20", "text":
        "What appears to be the primary, appropriate use for the model?",
     "type": "single_select",
     "options": ["Risk Assessment", "Preparedness Planning / CWPP Planning",
                 "Mitigation Planning/Resource Allocation",
                 "Evacuation Planning", "Readiness for Incident Response",
                 "Initial Attack during Response",
                 "Full Response Operations", "Post-Fire Recovery"],
     "narrative": True, "xlsx_row": 57},

    {"id": "q20a", "text":
        "If a Secondary use is appropriate, indicate which would be a "
        "viable secondary usage.",
     "type": "single_select",
     "options": ["Risk Assessment", "Preparedness Planning / CWPP Planning",
                 "Mitigation Planning/Resource Allocation",
                 "Evacuation Planning", "Readiness for Incident Response",
                 "Initial Attack during Response",
                 "Full Response Operations", "Post-Fire Recovery"],
     "narrative": True, "xlsx_row": 58},

    {"id": "q21", "text": "What are the greatest strengths of the model?",
     "type": "narrative", "options": None, "narrative": True, "xlsx_row": 59},

    {"id": "q22", "text":
        "What is the opportunity to ensemble this model with other models "
        "for a more accurate view of risk?",
     "type": "narrative", "options": None, "narrative": True, "xlsx_row": 60},

    # ---- Web-form-only questions (not in the Excel rubric) ----
    {"id": "qf1", "source": "science_form",
     "form_question_name": "indicate_the_model_organization",
     "text": 'Indicate the Model Organization Name',
     "type": "single_select",
     "options": [
         'A-T-S',
         'Envitel',
         'FireSafe Analytics',
         'FiSci',
         'FlameMapper',
         'French National Center for Scientific Research',
         'IMAGECAT',
         'Minerva University',
         'Opal AI',
         'OroraTech',
         'Pinepeak',
         'San Jose State University',
         'SkyTL',
         'Spatial Informatics Group',
         'Tenax AI',
         'UC Berkley / ElmFire',
         'UC Irvine',
         'UC Los Angeles',
         'Univ of Santiago, Chile',
         'Univ of Nevada',
         'Vanderbilt University',
         'Vibrant Planet',
         'XyloPlan',
         'Zesty AI',
     ],
     "narrative": False,
     "xlsx_row": None},
    {"id": "qf2", "source": "science_form",
     "form_question_name": "on_a_scale_of_1_5_did_the_model",
     "text":
         "On a scale of 1-5, did the model's use of building characteristics "
         'provide insights on wildfire risk in the urban/suburban area?',
     "type": "single_select",
     "options": [
         'Strongly disagree',
         'Disagree',
         'Neutral',
         'Agree',
         'Strongly agree',
     ],
     "narrative": False,
     "team_only": True,  # subjective rating / reviewer experience — skip extraction
     "xlsx_row": None},
    {"id": "qf3", "source": "science_form",
     "form_question_name": "since_the_model_does_not_provid",
     "text":
         'Since the model does not provide "Risk", indicate which of the '
         'following measures the model results provide.',
     "type": "multi_select",
     "options": [
         'Hazard Exposure',
         'Vulnerability',
         'Likelihood',
         'Burn Probability',
         'None',
     ],
     "narrative": False,
     "xlsx_row": None},
    {"id": "qf4", "source": "science_form",
     "form_question_name": "describe_how_the_model_expresse",
     "text": 'Describe how the model expresses risk in economic or financial terms.',
     "type": "narrative",
     "options": None,
     "narrative": True,
     "xlsx_row": None},
    {"id": "qf5", "source": "science_form",
     "form_question_name": "on_a_scale_rate_the_effectivene",
     "text":
         "On a scale, rate the effectiveness of the model's resulting risk rank, "
         'score, or value.',
     "type": "single_select",
     "options": [
         'Strongly disagree',
         'Disagree',
         'Neutral',
         'Agree',
         'Strongly agree',
     ],
     "narrative": False,
     "team_only": True,  # subjective rating / reviewer experience — skip extraction
     "xlsx_row": None},
    {"id": "qf6", "source": "science_form",
     "form_question_name": "describe_any_gaps_or_other_resu",
     "text":
         'Describe any gaps or other results that did not appear to align with '
         'the area.',
     "type": "narrative",
     "options": None,
     "narrative": True,
     "xlsx_row": None},
]


# Practitioner Eval Rubric. The workbook has TWO rows numbered "2" (an
# evaluator-ID question and a both-locations question). We disambiguate
# the second one by reading it as q2b, even though the rubric's "No"
# column labels it as 2. Render order in the xlsx is preserved by the
# explicit xlsx_row field.
PRACTITIONER_RUBRIC = [
    {"id": "q1", "text": "Model Name",
     "type": "text", "options": None, "narrative": True, "xlsx_row": 2},

    {"id": "q2", "text": "Indicate your Evaluator ID",
     "type": "text", "options": None, "narrative": False, "xlsx_row": 3},

    {"id": "q2b", "text":
        "Did you review the model outputs for both \"Town of Forests\" and "
        "\"Town of Prairies\" before completing this evaluation?",
     "type": "single_select", "options": ["Yes", "No"],
     "narrative": False, "xlsx_row": 4,
     "_xlsx_id_label": "2"},

    {"id": "q3", "text":
        "What is the primary use of this model? Select only one based on "
        "your interpretation of the model submission.",
     "type": "single_select",
     "options": ["Risk Assessment", "Preparedness Planning / CWPP Planning",
                 "Mitigation Planning/Resource Allocation",
                 "Evacuation Planning", "Readiness for Incident Response",
                 "Initial Attack during Response",
                 "Full Response Operations", "Post-Fire Recovery"],
     "narrative": True, "xlsx_row": 5},

    {"id": "q4", "text":
        "If the model has a secondary use, please indicate what that is. "
        "Select only one based on your interpretation of the model submission.",
     "type": "single_select",
     "options": ["Risk Assessment", "Preparedness Planning / CWPP Planning",
                 "Mitigation Planning/Resource Allocation",
                 "Evacuation Planning", "Readiness for Incident Response",
                 "Initial Attack during Response",
                 "Full Response Operations", "Post-Fire Recovery"],
     "narrative": True, "xlsx_row": 6},

    {"id": "q4a", "text":
        "For models with pre-fire uses - Does the model clearly and directly "
        "inform risk reduction decisions?",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": False, "xlsx_row": 7},

    {"id": "q4b", "text":
        "If Yes - Indicate ONE primary decision point that the model clearly "
        "informs. Describe in narrative.",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q4a": "Yes"}, "xlsx_row": 8},

    {"id": "q4c", "text":
        "If Yes - Indicate a SECONDARY decision point that the model clearly "
        "informs. Describe in narrative.",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q4a": "Yes"}, "xlsx_row": 9},

    {"id": "q5", "text":
        "Does the model distinguish between environment types/classifications "
        "for wildland, wildland urban interface, wildland urban intermix, "
        "urban/suburban area?",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": False, "xlsx_row": 10},

    {"id": "q5a", "text":
        "If YES - Describe how the model accounts for the distinction in "
        "these environment types and how they are defined and measured in "
        "the area.",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q5": "Yes"}, "xlsx_row": 11},

    {"id": "q5b", "text":
        "If NO - Describe what environment type is applied to the area "
        "exclusively.",
     "type": "single_select",
     "options": ["Wildland", "Wildland Urban Interface",
                 "Wildland Urban Intermix", "Urban/Suburban", "Other"],
     "narrative": True, "branch_when": {"q5": "No"}, "xlsx_row": 12},

    {"id": "q6", "text":
        "Which types of information does the model provide and/or include? "
        "Select all that apply.",
     "type": "multi_select",
     "options": ["Fire Behavior in Wildland", "Fire Behavior in WUI",
                 "Fire Behavior in Urban/Built Environment"],
     "narrative": False, "xlsx_row": 13},

    {"id": "q6a", "text":
        "What modes of fire spread does the model appear to account for? "
        "Select all that apply.",
     "type": "multi_select",
     "options": ["Flame Spread", "Radiative Heat Transfer",
                 "Ember Cast/Ember Spread", "Other"],
     "narrative": False, "xlsx_row": 14},

    {"id": "q6b", "text":
        "If the model goes beyond wildfire behavior and spread, which of "
        "the following does it account for? Select all that apply.",
     "type": "multi_select",
     "options": ["Buildings / Structures", "Vulnerable Populations",
                 "Critical Infrastructure", "Transportation Networks",
                 "Water Resources", "Other"],
     "narrative": False, "xlsx_row": 15},

    {"id": "q6c", "text":
        "Which of the following information types does the model provide "
        "and/or include? Select all that apply.",
     "type": "multi_select",
     "options": ["Structure Impact/Damage Level", "Structure Loss",
                 "Ignition Likelihood", "Critical Infrastructure Impacts",
                 "Ecosystem/Environmental Impacts", "Economic Impacts",
                 "Other Community Impacts", "Other"],
     "narrative": False, "xlsx_row": 16},

    {"id": "q6d", "text":
        "Which type of information does the model most effectively provide? "
        "Describe in the narrative.",
     "type": "narrative", "options": None, "narrative": True, "xlsx_row": 17},

    {"id": "q7", "text":
        "Does the model provide parcel-, neighborhood-, or community-level "
        "insights detailed enough to guide mitigation actions and investments?",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "xlsx_row": 18},

    {"id": "q7a", "text":
        "If Yes above, which types of mitigation actions and investments "
        "does the model inform? Select all that apply.",
     "type": "multi_select",
     "options": ["Structural Hardening", "Neighborhood prioritization",
                 "Defensible Space mitigation",
                 "Code adoption and enforcement",
                 "Neighborhood Chipper/Fuels reduction", "Prescribed Burn",
                 "Other Landscape Fuels Reductions", "Other"],
     "narrative": False, "branch_when": {"q7": "Yes"}, "xlsx_row": 19},

    {"id": "q7b", "text":
        "Which of the following inputs does the model most clearly include "
        "and provide insights on?",
     "type": "multi_select",
     "options": ["Landscape Fuels", "Structure-level Resistance",
                 "Structure-level Susceptibility", "Local Topography",
                 "Fire Weather Risks", "Critical Infrastructure Resistance",
                 "Critical Infrastructure Vulnerabilities",
                 "Response and Fire Management/Suppression Capacity"],
     "narrative": False, "xlsx_row": 20},

    {"id": "q8", "text":
        "Is the model output classified as \"static\" or \"dynamic\"? "
        "Select one and provide notes in the narrative.",
     "type": "single_select", "options": ["Static", "Dynamic", "Unsure"],
     "narrative": True, "xlsx_row": 21},

    {"id": "q8a", "text":
        "IF DYNAMIC - Can users \"test\" mitigation strategies and compare "
        "outcomes such as implementation of specific fuels reductions, "
        "home hardening, or land-use changes for a specific location in "
        "different scenarios?",
     "type": "single_select", "options": ["Yes", "No", "Unsure"],
     "narrative": False, "branch_when": {"q8": "Dynamic"}, "xlsx_row": 22},

    {"id": "q9", "text":
        "Are the assumptions, limitations, and confidence level of the model "
        "output clearly conveyed to the user? Select one and describe in "
        "the narrative.",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "xlsx_row": 23},

    {"id": "q9a", "text": "Is the model easy to interpret?",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "xlsx_row": 24},

    {"id": "q9b", "text":
        "Does the model appear to be applied consistently across "
        "locations/jurisdictions while adaptive to local conditions? Select "
        "one and describe in the narrative.",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "xlsx_row": 25},

    {"id": "q9c", "text":
        "Does the model help identify disproportionately at-risk or "
        "vulnerable populations and account for these population variables "
        "in providing decision support for mitigation investments and/or "
        "response operations? Select one and describe in the narrative.",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "xlsx_row": 26},

    {"id": "q9d", "text":
        "Does the model quantify avoided losses, return on mitigation "
        "investment, and/or resilience benefits? Select one and describe "
        "which and how in the narrative.",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "xlsx_row": 27},

    {"id": "q9e", "text":
        "Does the model output include some kind of quantitative spatial "
        "distribution of risk (or it's equivalent)? Select all that applies "
        "and provide notes in the narrative.",
     "type": "multi_select",
     "options": ["Areas Impacted by Spread",
                 "Number of Structures Impacted", "Types of Structures",
                 "Specific Areas within a Subdivision",
                 "Specific Areas in the Wildland or WUI",
                 "Other (if other, list)"],
     "narrative": True, "xlsx_row": 28},

    {"id": "q9f", "text":
        "Does the model output provide a rank, score, or value for \"Risk\"?",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": False, "xlsx_row": 29},

    {"id": "q9g", "text":
        "If YES - Indicate the type of risk ranking or score provided.",
     "type": "single_select",
     "options": ["Binary", "Numerical", "Categorical", "Other"],
     "narrative": True, "branch_when": {"q9f": "Yes"}, "xlsx_row": 30,
     "_option_descriptions": {
         "Binary":      "Classify areas as either high risk or not. The risk score indicates whether an area exceeds a threshold.",
         "Numerical":   "Continuous risk scores that can be visualized in quantized buckets, even buckets, custom buckets.",
         "Categorical": "Assign discrete risk categories to areas. The risk score is a numeric identifier that maps to a category.",
     }},

    {"id": "q9h", "text":
        "If YES - Does the risk ranking or score include economic variables "
        "such as expected financial loss or value of loss avoided? If Yes, "
        "provide notes in the narrative.",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "branch_when": {"q9f": "Yes"}, "xlsx_row": 31},

    {"id": "q9i", "text":
        "Describe the risk ranking or score provided by the model based on "
        "your assessment.",
     "type": "narrative", "options": None, "narrative": True, "xlsx_row": 32},

    {"id": "q9j", "text":
        "Were there any risk scores, rankings, or values that did not appear "
        "to align with the characteristics of the area? Select one option "
        "and provide notes in the narrative.",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "xlsx_row": 33},

    {"id": "q9k", "text":
        "Were there any gaps in the risk score, rankings, or values for "
        "certain areas? Select one option and provide notes in the narrative.",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "xlsx_row": 34},

    {"id": "q9l", "text":
        "Is the model output a full measure or quantification of \"Risk\"?",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": False, "xlsx_row": 35},

    {"id": "q10", "text": "Describe the model's top 3 strengths.",
     "type": "narrative", "options": None, "narrative": True, "xlsx_row": 36},

    {"id": "q10a", "text": "What are the greatest strengths of the model?",
     "type": "narrative", "options": None, "narrative": True, "xlsx_row": 37},

    {"id": "q10b", "text":
        "Do you see an option to combine this model output with another "
        "model for a more accurate view of risk?",
     "type": "single_select", "options": ["Yes", "No", "Somewhat"],
     "narrative": True, "xlsx_row": 38},

    {"id": "q10c", "text": "If Yes - describe in the narrative.",
     "type": "narrative", "options": None, "narrative": True,
     "branch_when": {"q10b": "Yes"}, "xlsx_row": 39},

    # ---- Web-form-only questions (not in the Excel rubric) ----
    {"id": "qf1", "source": "practitioner_form",
     "form_question_name": "indicate_the_model_organization",
     "text": 'Indicate the Model Organization Name',
     "type": "single_select",
     "options": [
         'A-T-S',
         'Envitel',
         'FireSafe Analytics',
         'FiSci',
         'FlameMapper',
         'French National Center for Scientific Research',
         'IMAGECAT',
         'Minerva University',
         'Opal AI',
         'OroraTech',
         'Pinepeak',
         'San Jose State University',
         'SkyTL',
         'Spatial Informatics Group',
         'Tenax AI',
         'UC Berkley / ElmFire',
         'UC Irvine',
         'UC Los Angeles',
         'Univ of Santiago, Chile',
         'Univ of Nevada',
         'Vanderbilt University',
         'Vibrant Planet',
         'XyloPlan',
         'Zesty AI',
     ],
     "narrative": False,
     "xlsx_row": None},
    {"id": "qf2", "source": "practitioner_form",
     "form_question_name": "what_appears_to_be_a_third_appr",
     "text": 'What appears to be a third, appropriate use for the model?',
     "type": "single_select",
     "options": [
         'Risk Assessment',
         'Preparedness Planning (i.e. CWPP)',
         'Mitigation Planning/Prioritization',
         'Evacuation Planning',
         'Readiness for Incident Response',
         'Initial Attack Phase',
         'Full Response Operations',
         'Post-Fire Recovery',
         'NONE',
     ],
     "narrative": False,
     "xlsx_row": None},
    {"id": "qf3", "source": "practitioner_form",
     "form_question_name": "which_types_of_mitigation_actio",
     "text":
         'Which types of mitigation actions and investments does the model '
         'inform?',
     "type": "multi_select",
     "options": [
         'Parcel-level Structural Hardening',
         'Parcel-level Defensible Space Mgmt',
         'Neighborhood-level Structural Hardening',
         'Neighborhood-level Defensible Space',
         'Parcel-level Prioritization',
         'Neighborhood-level Prioritization',
         'Code Adoption Decisions',
         'Code Enforcement Strategies & Prioritization',
         'Evacuation Preparedness Education',
         'Site Selection for Temporary Fire Refuge Areas (TFRAs)',
         'Neighborhood-level Fuels Reduction',
         'Wildland/WUI Landscape Fuels Reduction',
         'Fuels Reduction Prioritization',
         'NONE',
     ],
     "narrative": False,
     "xlsx_row": None},
    {"id": "qf4", "source": "practitioner_form",
     "form_question_name": "field_88",
     "text":
         'Does the model clearly and directly inform response operations '
         'decisions?',
     "type": "single_select",
     "options": [
         'Yes',
         'No',
     ],
     "narrative": False,
     "xlsx_row": None},
    {"id": "qf5", "source": "practitioner_form",
     "form_question_name": "which_types_of_response_functio",
     "text": 'Which types of response functions does the model clearly support?',
     "type": "multi_select",
     "options": [
         'Initial Attack Actions & Tactics',
         'Ground Suppression Resource Positioning and Tasking',
         'Aviation Suppression Resource Positioning and Tasking',
         'Determining Suppression Material Types and Volumes',
         'Determining Timing for Suppression',
         'Tasking Incident Management Teams',
         'Locating Staging Locations',
         'Locating Responder Safety Zones',
     ],
     "narrative": False,
     "xlsx_row": None},
    {"id": "qf6", "source": "practitioner_form",
     "form_question_name": "which_of_the_following_environm",
     "text":
         'Which of the following environment types does the model appear to most '
         'effectively model?',
     "type": "single_select",
     "options": [
         'Wildland',
         'Wildland Urban Interface (WUI)',
         'Wildland Urban Intermix',
         'Urban/Suburban',
     ],
     "narrative": False,
     "xlsx_row": None},
    {"id": "qf7", "source": "practitioner_form",
     "form_question_name": "can_you_discern_whether_its_app",
     "text":
         "Can you discern whether it's applying different fire behavior "
         'variables between vegetation and buildings?',
     "type": "single_select",
     "options": [
         'Yes',
         'No',
     ],
     "narrative": False,
     "xlsx_row": None},
    {"id": "qf8", "source": "practitioner_form",
     "form_question_name": "indicate_which_of_the_following",
     "text": 'Indicate which of the following insights the model provides for.',
     "type": "multi_select",
     "options": [
         'Specific Structures at Varying Risk Levels',
         'Specific Structures deemed "Super Spreaders"',
         'Neighborhoods at Varying Risk Levels',
         'High risk vegetation pathways in WUI',
         'High risk areas in Wildland',
         'Estimated ROI for specific fuels reduction',
         'Estimated ROI for specific structural hardening',
         'None',
     ],
     "narrative": False,
     "xlsx_row": None},
    {"id": "qf9", "source": "practitioner_form",
     "form_question_name": "describe_your_experience_review",
     "text":
         'Describe your experience reviewing, interpreting, and understanding '
         "the model's results.",
     "type": "narrative",
     "options": None,
     "narrative": True,
     "team_only": True,  # subjective rating / reviewer experience — skip extraction
     "xlsx_row": None},
    {"id": "qf10", "source": "practitioner_form",
     "form_question_name": "describe_how_the_model_provides",
     "text": 'Describe how the model provides these insights.',
     "type": "narrative",
     "options": None,
     "narrative": True,
     "xlsx_row": None},
    {"id": "qf11", "source": "practitioner_form",
     "form_question_name": "describe_how_the_model_expresse",
     "text": 'Describe how the model expresses risk in economic or financial terms.',
     "type": "narrative",
     "options": None,
     "narrative": True,
     "xlsx_row": None},
    {"id": "qf12", "source": "practitioner_form",
     "form_question_name": "on_a_scale_rate_the_effectivene",
     "text":
         "On a scale, rate the effectiveness of the model's resulting risk rank, "
         'score, or value.',
     "type": "single_select",
     "options": [
         'Strongly disagree',
         'Disagree',
         'Neutral',
         'Agree',
         'Strongly agree',
     ],
     "narrative": False,
     "team_only": True,  # subjective rating / reviewer experience — skip extraction
     "xlsx_row": None},
    {"id": "qf13", "source": "practitioner_form",
     "form_question_name": "describe_any_gaps_or_other_resu",
     "text":
         'Describe any gaps or other results that did not appear to align with '
         'the area.',
     "type": "narrative",
     "options": None,
     "narrative": True,
     "xlsx_row": None},
    {"id": "qf14", "source": "practitioner_form",
     "form_question_name": "provide_any_relevant_notes_and",
     "text": 'Provide any relevant notes and observations on the model results.',
     "type": "narrative",
     "options": None,
     "narrative": True,
     "xlsx_row": None},
    {"id": "qf15", "source": "practitioner_form",
     "form_question_name": "what_types_of_additional_insigh",
     "text":
         'What types of additional insights and/or other decisions would the '
         'model help support IF it were combined with another model?',
     "type": "narrative",
     "options": None,
     "narrative": True,
     "xlsx_row": None},
]


# Map a public skill name to its schema + sheet name. Order matters for
# error messages, not for storage.
SCHEMAS = {
    "science":      {"schema": SCIENCE_RUBRIC,      "sheet": SCIENCE_SHEET},
    "practitioner": {"schema": PRACTITIONER_RUBRIC, "sheet": PRACTITIONER_SHEET},
}


def get_schema(rubric_type: str):
    if rubric_type not in SCHEMAS:
        raise ValueError(
            f"Unknown rubric_type {rubric_type!r}. "
            f"Must be one of {list(SCHEMAS)}.")
    return SCHEMAS[rubric_type]["schema"]


def schema_by_id(rubric_type: str):
    return {q["id"]: q for q in get_schema(rubric_type)}


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

class RubricValidationError(ValueError):
    """Raised when answers don't match the schema."""


def _validate_one(rubric_type: str, qid: str, answer):
    """Validate `answer` against the schema entry for `qid`."""
    if answer is None:
        return
    by_id = schema_by_id(rubric_type)
    if qid not in by_id:
        raise RubricValidationError(
            f"Unknown question id {qid!r} for rubric_type={rubric_type!r}")
    spec = by_id[qid]
    t = spec["type"]
    opts = spec.get("options")

    if t == "single_select":
        v = answer.get("value")
        if v is not None and opts is not None and v not in opts:
            raise RubricValidationError(
                f"{qid}: value {v!r} not in options {opts}")
    elif t == "multi_select":
        sel = answer.get("selected", [])
        if not isinstance(sel, list):
            raise RubricValidationError(
                f"{qid}: 'selected' must be a list, got {type(sel).__name__}")
        bad = [x for x in sel if opts is not None and x not in opts]
        if bad:
            raise RubricValidationError(
                f"{qid}: invalid selections {bad} (not in {opts})")
    elif t == "narrative":
        if "notes" not in answer:
            raise RubricValidationError(
                f"{qid}: narrative-only question requires a 'notes' field")
    # `text` has no constraints.


def validate_answers(rubric_type: str, answers: dict) -> None:
    for qid, ans in (answers or {}).items():
        _validate_one(rubric_type, qid, ans)


# ---------------------------------------------------------------------------
# Helpers the agent should USE rather than re-derive categorization
# ---------------------------------------------------------------------------

def extract_pdf_text(pdf_path, max_pages: int = 80) -> str:
    """Read a PDF and return its text content as a single string.

    Use this for the team's Final Report PDF and any model-documentation
    PDF in `additional_resources/`. Without this helper, the agent
    fell back to `read_file()` on the binary PDF and got garbage,
    causing it to flag PDF-derived questions as evidence gaps even
    though the source was locally available.

    Parameters
    ----------
    pdf_path : str | Path
        Path to the PDF file.
    max_pages : int
        Safety cap to avoid runaway parses on enormous PDFs. Default 80
        covers any team's Final Report or model documentation; the
        exercise's own Situation Manual is 38 pages.

    Returns
    -------
    str — concatenated page text. Each page is separated by a
    `--- PAGE N ---` marker so the agent can cite "page N" if useful.

    Raises ``FileNotFoundError`` if the path doesn't exist;
    ``RuntimeError`` if pypdf can't be imported (image is too old).
    """
    p = Path(pdf_path)
    if not p.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")
    try:
        from pypdf import PdfReader
    except ImportError as e:
        raise RuntimeError(
            f"pypdf is not available in this Sage image. "
            f"PDF parsing requires v1.1.25+ (built 2026-05-20 or later)."
        ) from e
    reader = PdfReader(str(p))
    pages = []
    n = min(len(reader.pages), max_pages)
    for i in range(n):
        try:
            txt = reader.pages[i].extract_text() or ""
        except Exception as e:  # noqa: BLE001
            txt = f"[page {i+1}: extraction failed: {type(e).__name__}: {e}]"
        pages.append(f"--- PAGE {i+1} ---\n{txt}")
    if len(reader.pages) > max_pages:
        pages.append(f"[... truncated; {len(reader.pages) - max_pages} additional pages not extracted ...]")
    return "\n\n".join(pages)


def cascade_skip(parent_id: str, child_ids: list[str], parent_note: str = "") -> dict:
    """Record children as cascade evidence gaps because the parent itself
    is an evidence gap (parent value is null because we lack evidence).

    Use when the parent question's answer is unknown — the children are
    unreachable but might be relevant if the parent were answered. The
    correct dict for this is `evidence_gaps`, NOT `branching_notes`.

    Returns
    -------
    A dict suitable to merge into `evidence_gaps`. Example::

        evidence_gaps["q4"]  = "Question: '... static or dynamic?' — Final Report pending."
        evidence_gaps.update(cascade_skip(
            "q4", ["q4a", "q4b", "q4c", "q4d", "q4e"],
            parent_note="Final Report pending — model type unknown",
        ))
    """
    note = parent_note or "parent is an evidence gap"
    return {cid: f"cascade — parent {parent_id} is an evidence gap ({note})"
            for cid in child_ids}


def branch_skip(parent_id: str, parent_value, child_ids) -> dict:
    """Record children as branching_notes because the parent's actual
    answer excludes them (per the rubric's branching rules).

    Use only when the parent question has an answer that the rubric's
    branching rules use to exclude these children. NOT for "parent is
    null" — use `cascade_skip()` for that.

    Returns
    -------
    A dict suitable to merge into `branching_notes`. Example::

        # q5 was answered "No" → q5a is definitively n/a
        branching_notes.update(branch_skip("q5", "No", ["q5a"]))

        # q4 was answered "Static" → q4a, q4b, q4c, q4d don't apply
        branching_notes.update(branch_skip("q4", "Static",
                                           ["q4a", "q4b", "q4c", "q4d"]))
    """
    # Defend against the common mistake of passing a string instead of a
    # list of question ids. `for cid in "If STATIC..."` would iterate
    # character-by-character and build a dict keyed by individual letters
    # — a real bug we hit in 2026-05-23 rubric runs.
    if isinstance(child_ids, str):
        raise TypeError(
            f"branch_skip(child_ids=...) must be a list/tuple of question ids, "
            f"not a string. Got: {child_ids!r}. "
            f"Did you mean ['{child_ids}'] (single-element list)?"
        )
    return {cid: f"not applicable — {parent_id} = {parent_value!r}"
            for cid in child_ids}


# ---------------------------------------------------------------------------
# Filenames
# ---------------------------------------------------------------------------

def _sanitize_for_filename(name: str) -> str:
    if not name:
        return "submission"
    name = re.sub(r'[\\/*?:"<>|]', "_", name)
    name = re.sub(r"\s+", "_", name).strip("._-")
    return name or "submission"


def review_json_path(ws_root, rubric_type: str) -> Path:
    """Path the rubric JSON will be written to."""
    return Path(ws_root) / f"_rubric_{rubric_type}.json"


def review_xlsx_path(ws_root, workspace_name: str | None, rubric_type: str) -> Path:
    """Path the filled xlsx will be written to."""
    ws_root = Path(ws_root)
    ws = workspace_name or ws_root.name
    return ws_root / f"_rubric_{rubric_type}_{_sanitize_for_filename(ws)}.xlsx"


def review_already_exists(ws_root, rubric_type: str) -> bool:
    return review_json_path(ws_root, rubric_type).exists()


# ---------------------------------------------------------------------------
# Save JSON + render xlsx
# ---------------------------------------------------------------------------

def _check_coverage(rubric_type: str, answers: dict, branching_notes: dict,
                    evidence_gaps: dict, team_only: dict | None = None) -> list[str]:
    """Return question ids that are unanswered AND missing from any notes dict.

    An unanswered question (value is None) must have a recorded reason in
    one of `branching_notes` (parent excludes it), `evidence_gaps`
    (couldn't determine from available sources), or `team_only` (Sage
    deliberately skipped extraction because the question asks for the
    reviewer's subjective experience). Silent nulls confuse downstream
    consumers — a reviewer sees no answer and no explanation and doesn't
    know what to do.
    """
    schema_ids = {q["id"] for q in get_schema(rubric_type)}
    nulls = {qid for qid in schema_ids
             if qid in answers and answers[qid] is None}
    covered = (set(branching_notes or {})
               | set(evidence_gaps or {})
               | set(team_only or {}))
    missing = sorted(nulls - covered)
    return missing


def team_only_default_notes(rubric_type: str) -> dict:
    """Return a ready-made team_only dict for the rubric, marking each
    schema entry whose `team_only` flag is True with a standard note.
    Saves the agent from having to enumerate these manually."""
    out = {}
    for spec in get_schema(rubric_type):
        if spec.get("team_only"):
            out[spec["id"]] = (
                "Team-rated question — the review agent does not extract "
                "this from submission materials. The team's web-form answer "
                "is the only source for this question."
            )
    return out


def write_rubric_review(
    ws_root,
    rubric_type: str,
    answers: dict,
    branching_notes: dict | None = None,
    evidence_gaps: dict | None = None,
    team_only: dict | None = None,
    model_name_for_q1: str | None = None,
    write_xlsx: bool = True,
) -> Path:
    """Save the rubric JSON and auto-render the filled xlsx alongside it.

    Parameters
    ----------
    ws_root : str | Path
        The workspace folder (must contain _manifest.json from the prior
        download step).
    rubric_type : str
        Either "science" or "practitioner".
    answers : dict
        Map from question id (e.g., "q3", "q9bi") to a dict in one of these
        shapes:
          single_select  : {"value": <option>, "notes": "..."}    (notes optional)
          multi_select   : {"selected": [<options>], "notes": "..."}
          narrative      : {"notes": "..."}
          text           : {"value": "<string>", "notes": "..."}   (notes optional)
        Omit keys that don't apply due to branching or missing evidence;
        record the reason in branching_notes or evidence_gaps.
    branching_notes : dict, optional
        Map from question id to a one-line reason it was SKIPPED by the
        rubric's branching rules.
    evidence_gaps : dict, optional
        Map from question id to a one-line reason it COULD have been
        answered, but the agent lacked source material.
    team_only : dict, optional
        Map from question id to a one-line reason Sage deliberately
        skipped extraction (subjective rating / reviewer experience).
        If omitted, this is auto-populated from the schema's `team_only`
        flag via `team_only_default_notes()`.
    write_xlsx : bool, optional (default True)
        When False, skip the auto-rendered filled xlsx and write only the
        JSON. Used by the combined `wildfire-rubric-review` skill which
        produces JSON-only output (the boss's web forms have replaced the
        filled-xlsx-upload reviewer workflow).
    model_name_for_q1 : str, optional
        Convenience — if provided and q1 is not in answers, q1's value
        defaults to this.

    Returns
    -------
    Path of the JSON file. The xlsx is written next to it; the .xlsx
    path is available via review_xlsx_path().
    """
    ws_root = Path(ws_root)
    if not (ws_root / "_manifest.json").exists():
        raise ValueError(
            f"{ws_root} doesn't look like a downloaded workspace "
            "(no _manifest.json found).")
    schema = get_schema(rubric_type)

    answers = dict(answers or {})
    if model_name_for_q1 and "q1" not in answers:
        answers["q1"] = {"value": model_name_for_q1, "notes": ""}

    validate_answers(rubric_type, answers)

    record_answers = {}
    for spec in schema:
        qid = spec["id"]
        record_answers[qid] = answers.get(qid)

    branching_notes = branching_notes or {}
    evidence_gaps = evidence_gaps or {}
    # Defend against caller passing non-string values (e.g., a dict-of-chars
    # produced when branch_skip(child_ids=<a string>) is misused).
    for bucket_name, bucket in [("branching_notes", branching_notes),
                                  ("evidence_gaps", evidence_gaps)]:
        bad = [(k, type(v).__name__) for k, v in bucket.items()
               if not isinstance(v, str)]
        if bad:
            raise RubricValidationError(
                f"{bucket_name} values must be strings. Got non-string for: "
                f"{bad}. Each value should be a one-line reason. If you used "
                f"branch_skip(), make sure child_ids is a list like "
                f"['q4e'], NOT the question text string."
            )
    # Auto-populate team_only from the schema. Caller can override by
    # passing an explicit dict (e.g., to add per-question notes).
    if team_only is None:
        team_only = team_only_default_notes(rubric_type)

    # Coverage check: every null answer MUST be explained in one of the
    # three notes dicts. We RAISE rather than just warn — past runs showed
    # the agent ignores printed warnings and reports success anyway. A
    # raise forces the agent to fix the gap and re-run.
    missing = _check_coverage(rubric_type, record_answers,
                              branching_notes, evidence_gaps, team_only)
    if missing:
        raise RubricValidationError(
            f"Coverage incomplete: {len(missing)} question id(s) are null "
            f"in 'answers' but missing from ALL THREE notes dicts: {missing}.\n\n"
            f"Each must be placed in exactly one of four buckets:\n"
            f"  1. answers[qid] = {{...}}  — if the question is answerable\n"
            f"  2. branching_notes[qid]    — if the parent has a real value "
            f"that excludes this child (use branch_skip())\n"
            f"  3. evidence_gaps[qid]      — if the question requires a "
            f"source the agent doesn't have, OR the parent is itself an "
            f"evidence gap (use cascade_skip() for the second case)\n"
            f"  4. team_only[qid]          — if the schema marks it "
            f"team_only=True (subjective rating / reviewer experience). "
            f"This bucket is auto-populated; if you see entries here in "
            f"the missing list, the schema is out of sync.\n\n"
            f"Fix the wrapper script and re-run. The JSON was NOT written."
        )

    out = {
        "workspace_name": ws_root.name,
        "rubric_type": rubric_type,
        "reviewed_at": datetime.now(timezone.utc).isoformat(),
        "answers": record_answers,
        "branching_notes": branching_notes,
        "evidence_gaps": evidence_gaps,
        "team_only": team_only,
    }
    # Surface the coverage warning inside the JSON too, so the xlsx
    # render can mark those rows distinctly and downstream synthesis
    # can flag the workspace.
    if missing:
        out["coverage_warnings"] = {
            "missing_explanations": missing,
            "note": (
                "These question ids are null in 'answers' but absent from "
                "both 'branching_notes' and 'evidence_gaps'. They should be "
                "filled in a re-run."
            ),
        }

    json_path = review_json_path(ws_root, rubric_type)
    json_path.write_text(json.dumps(out, indent=2))

    # Optionally auto-render the filled xlsx. Render failures are
    # warnings, not errors — the JSON is the source of truth.
    if write_xlsx:
        xlsx_path = review_xlsx_path(ws_root, ws_root.name, rubric_type)
        try:
            render_xlsx(rubric_type, out, xlsx_path)
        except Exception as e:  # noqa: BLE001
            print(f"Warning: xlsx render failed ({type(e).__name__}: {e}). "
                  f"JSON written successfully to {json_path}.")

    return json_path


def render_xlsx(rubric_type: str, record: dict, xlsx_out_path: Path) -> Path:
    """Render `record` as a filled copy of the blank rubric template.

    The rubric template (`Exercise-Evaluation-Rubric-20260518.xlsx`)
    contains both the Science and Practitioner sheets. We copy the
    template wholesale, then overwrite cells on the target sheet only;
    the other sheet remains blank so the reviewer can still see the
    other rubric's structure if they need to.

    For each question row, we write:
      - single_select  : the chosen option label into the appropriate
                         Response Option column (and a "✓" prefix).
      - multi_select   : "✓" markers into every column whose header
                         option was selected.
      - narrative      : the notes string into the Open Text Narrative column.
      - text           : the value into Response Option A (the rubric's
                         "Dropdown Menu" cell), narrative if present.
      - branched-out   : the Narrative column gets "(not applicable — <reason>)"
      - evidence-gap   : the Narrative column gets "(pending — <reason>)"
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        import subprocess, sys
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "-q", "--user", "openpyxl"]
        )
        import site
        site.main()
        from openpyxl import load_workbook

    if not RUBRIC_TEMPLATE_XLSX.exists():
        raise FileNotFoundError(
            f"Rubric template missing at {RUBRIC_TEMPLATE_XLSX}. "
            "The xlsx must ship next to wildfire_rubric.py.")

    xlsx_out_path = Path(xlsx_out_path)
    xlsx_out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(RUBRIC_TEMPLATE_XLSX, xlsx_out_path)

    wb = load_workbook(xlsx_out_path)
    sheet_name = SCHEMAS[rubric_type]["sheet"]
    ws = wb[sheet_name]

    answers = record.get("answers", {}) or {}
    branching = record.get("branching_notes", {}) or {}
    gaps = record.get("evidence_gaps", {}) or {}
    missing_coverage = set(
        (record.get("coverage_warnings") or {}).get("missing_explanations") or []
    )

    # Column layout: A=No, B=Question, C..J=Response Option A..H,
    # K=Open Text Narrative. (11 columns total.)
    OPTION_COLS = ["C", "D", "E", "F", "G", "H", "I", "J"]
    NARRATIVE_COL = "K"

    schema = get_schema(rubric_type)
    for spec in schema:
        qid = spec["id"]
        row = spec["xlsx_row"]
        # Web-form-only entries have no row in the Excel template — skip
        # them. They appear in the JSON output but not the rendered xlsx.
        if row is None:
            continue
        t = spec["type"]
        opts = spec.get("options")

        ans = answers.get(qid)

        # Branched-out or evidence-gap: note in the narrative column
        if ans is None:
            note = None
            if qid in branching:
                note = f"(not applicable — {branching[qid]})"
            elif qid in gaps:
                note = f"(pending — {gaps[qid]})"
            elif qid in missing_coverage:
                # Coverage-warning rows: agent left this null without a
                # reason. Mark them distinctly so the reviewer sees they
                # need attention.
                note = "(⚠ MISSING — unanswered with no reason recorded; re-run skill to patch)"
            if note:
                ws[f"{NARRATIVE_COL}{row}"] = note
            continue

        # Active answer — fill the option / multi-select columns and narrative
        if t == "single_select":
            v = ans.get("value")
            if v is not None and opts is not None and v in opts:
                col = OPTION_COLS[opts.index(v)]
                ws[f"{col}{row}"] = f"✓ {v}"
            notes = ans.get("notes")
            if notes:
                ws[f"{NARRATIVE_COL}{row}"] = notes

        elif t == "multi_select":
            sel = set(ans.get("selected") or [])
            if opts:
                for i, option in enumerate(opts):
                    if option in sel:
                        ws[f"{OPTION_COLS[i]}{row}"] = f"✓ {option}"
            notes = ans.get("notes")
            if notes:
                ws[f"{NARRATIVE_COL}{row}"] = notes

        elif t == "narrative":
            notes = ans.get("notes")
            if notes:
                ws[f"{NARRATIVE_COL}{row}"] = notes

        elif t == "text":
            v = ans.get("value")
            if v is not None:
                ws[f"{OPTION_COLS[0]}{row}"] = v
            notes = ans.get("notes")
            if notes:
                ws[f"{NARRATIVE_COL}{row}"] = notes

    wb.save(xlsx_out_path)
    return xlsx_out_path
